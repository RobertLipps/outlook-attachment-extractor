
# ------------------------------------------------------ IMPORTS -------------------------------------------------------

# ------------------------------ Standard Library Imports ------------------------------

import configparser
from datetime import date, datetime, timedelta
import fnmatch
import gc
import logging
import os
import re
import sys
from enum import Enum

# -------------------------------- Third-Party Imports ---------------------------------

import pandas as pd
import pythoncom
import pytz
from openpyxl import load_workbook
import win32com.client

# ------------------------------------------------- Logger Initiation --------------------------------------------------

# Configure logging globally
log_file_path = "default.log"  # Temporary default; will be overridden in main()
logging.basicConfig(filename=log_file_path, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ----------------------------------------------------- CONSTANTS ------------------------------------------------------

# Pulled from initialization file
config = configparser.ConfigParser()
config.read('Restatement_Process_Config.ini')

def get_config_value(section, key):
    value = config.get(section, key, fallback=None)
    if not value:
        raise ValueError(f"Missing required config value: [{section}] {key}")
    return value


# Excel File & Sheet Names
EXCEL_FILE = get_config_value('EXCEL', 'excel_file')
AUTOMATED_SHEET = get_config_value('EXCEL', 'automated_sheet')
SETTINGS_SHEET = get_config_value('EXCEL', 'settings_sheet')
REVISION_SHEET = get_config_value('EXCEL', 'revisions_sheet')

# Outlook Mailbox & Folder
SHARED_MAILBOX = get_config_value('Outlook', 'shared_mailbox')
FOLDER_NAME = get_config_value('Outlook', 'folder_name')

#Directories
BASE_PATH = get_config_value('Directories', 'daily_imported_statements')

# ---------------------------------------------------- Enum types ------------------------------------------------------

class OutlookMAPIType(Enum):
    INBOX = 6
    MAIL = 43

class OutlookFolderName(Enum):
    INBOX = "inbox"
    SUBFOLDER = "subfolder"
    CUSTOM = "custom"

# ------------------------------------------------- Utility Functions --------------------------------------------------

def get_filter_time(reference_date, hour, minute, timezone="US/Eastern"):
    tz = pytz.timezone(timezone)
    naive_time = datetime.combine(reference_date, datetime.strptime(f"{hour}:{minute}", "%H:%M").time())

    return tz.localize(naive_time)


def get_prior_business_days(n):
    reference_date = date.today()

    if n == 0:
        while reference_date.weekday() >= 5:
            reference_date -= timedelta(days=1)
        return reference_date

    business_days_count = 0
    target_date = reference_date

    while business_days_count < abs(n):
        target_date -= timedelta(days=1)
        if target_date.weekday() < 5:
            business_days_count += 1

    return target_date


def make_import_archive_path(reference_date):
    base_path = BASE_PATH
    yearly_folder = f"{reference_date:%Y}"
    monthly_folder = f"{reference_date:%B}"
    daily_folder = f"COB {reference_date:%m.%d.%Y}"
    dir_path = os.path.join(base_path, yearly_folder, monthly_folder, daily_folder)

    os.makedirs(dir_path, exist_ok=True)

    return dir_path

# ----------------------------------------------------- Main Class -----------------------------------------------------

class RestatementProcessor:
    def __init__(self, excel_file,archive_folder, mailbox_name=None, folder_name=None):
        self.excel_file = excel_file
        self.wb = None
        self.auto_ws = None
        self.mailbox_name = mailbox_name
        self.archive_folder = archive_folder
        self.mapping_df = None
        self.mapping_dict = None
        self.folder = None
        self.items = None
        self.excel_doc_updates = None


    @staticmethod
    def _get_named_cell_value(wb, names=None):
        if names is None:
            names = ["CBD", "PBD", "P2BD", "Start_Time", "End_Time", "Execution_Time"]

        named_cells = {}

        for name in names:
            defined_name = wb.defined_names.get(name)
            if defined_name:
                sheet_name, cell_address = list(defined_name.destinations)[0]
                sheet = wb[sheet_name]
                cell = sheet[cell_address]
                named_cells[name] = cell
            else:
                logger.warning(f"Named range '{name}' not found.")
                named_cells[name] = None

        return named_cells


    def reset_excel_template(self, cbd, pbd, p2bd, script_start):
        try:
            wb = load_workbook(self.excel_file)
            self.wb = wb
            logger.info("Workbook loaded successfully.")
        except KeyError as e:
            logger.error(f"Cannot load Workbook: {e}")
            return

        try:
            auto_ws = wb[AUTOMATED_SHEET]
            self.auto_ws = auto_ws
        except KeyError as e:
            logger.error(f"Automated sheet not found: {e}")
            return

        named_cells = RestatementProcessor._get_named_cell_value(wb)

        named_cells["CBD"].value = cbd.strftime("%Y-%m-%d")
        named_cells["PBD"].value = pbd.strftime("%Y-%m-%d")
        named_cells["P2BD"].value = p2bd.strftime("%Y-%m-%d")
        named_cells["Start_Time"].value = script_start.strftime("%Y-%m-%d %H:%M:%S")

        logger.info(f"Business dates calculated: CBD={cbd}, PBD={pbd}, P2BD={p2bd}")

        row = 2
        while auto_ws[f"E{row}"].value is not None:
            for col in ["E", "F"]:
                auto_ws[f"{col}{row}"].value = None
            row += 1

        logger.info(f"Status reset completed. Cleared rows: {row - 2}")


    def build_dictionary_from_excel(self):
        try:
            df = pd.read_excel(self.excel_file, sheet_name=self.auto_ws.title, engine="openpyxl")
            df.columns = df.columns.str.strip().str.lower()
            self.mapping_df = df
        except Exception as e:
            logger.error(f"Failed to read workbook: {e}")
            return None, None

        # Validate required columns
        required_cols = ['sender', 'subject', 'attachment', 'savename', 'status']
        missing = [col for col in required_cols if col not in df.columns]
        if missing:
            raise KeyError(f"Missing required columns in mapping table: {', '.join(missing)}")

        mapping_dict = {}

        for idx, row in df.iterrows():
            key = (
                str(row['sender']).lower().strip(),
                str(row['subject']).lower().strip()
            )
            value = {
                'AttachmentPattern': str(row['attachment']).lower().strip(),
                'SaveName': str(row['savename']).strip(),
                'RowIndex': idx
            }
            mapping_dict.setdefault(key, []).append(value)
            self.mapping_dict = mapping_dict

        logger.info(f"Mapping dictionary created successfully: {len(mapping_dict)} unique keys.")

        return df, mapping_dict


    def connect_outlook(self, folder_name=None, folder_type=None):

        if folder_type is None:
            folder_type = OutlookFolderName.INBOX

        outlook = win32com.client.Dispatch("Outlook.Application")
        namespace = outlook.GetNamespace("MAPI")
        recipient = namespace.CreateRecipient(self.mailbox_name)

        logger.info(f"Connecting to Outlook mailbox: {self.mailbox_name}")

        if not recipient.Resolve():
            raise Exception(f"Could not resolve shared mailbox: {self.mailbox_name}")

        if folder_type in [OutlookFolderName.CUSTOM, OutlookFolderName.SUBFOLDER] and not folder_name:
            raise ValueError("folder_name must be provided for 'custom' or 'subfolder' folder types.")

        shared_inbox = namespace.GetSharedDefaultFolder(recipient, OutlookMAPIType.INBOX.value)

        try:
            if folder_type.value == "custom":
                folder = shared_inbox.Parent.Folders(folder_name)
            elif folder_type.value == "subfolder":
                folder = shared_inbox.Folders(folder_name)
            else:
                folder = shared_inbox
        except Exception as folder_error:
            raise Exception(f"Could not access folder '{folder_name}': {folder_error}")

        logger.info(f"Connected to folder: {folder.Name} (Type: {folder_type.value})")

        self.folder = folder
        return folder


    def get_items(self, get_filter_time_func):
        if self.folder is None:
            raise ValueError("Outlook folder not connected. Call connect_outlook() first.")

        filter_time = get_filter_time_func()

        filter_str = f"[ReceivedTime] >= '{filter_time.strftime('%m/%d/%Y %H:%M %p')}'"
        items = self.folder.Items.Restrict(filter_str)
        items.Sort("[ReceivedTime]", True)
        self.items = items

        logger.info(f"Retrieved {len(items)} messages from Outlook folder: {self.folder.Name}")

        return items


    def match_and_save_attachments(self):
        if self.items is None:
            raise ValueError("Folder items haven't been retrieved, call get_items first")

        updates = {}

        for i, message in enumerate(self.items):
            try:
                if message.Class == OutlookMAPIType.MAIL.value and message.Attachments.Count > 0:
                    sender = str(message.SenderEmailAddress).lower().strip()
                    subject = str(message.Subject).lower().strip()
                    key = (sender, subject)
                    possible_matches = self.mapping_dict.get(key, [])

                    logger.info(f"Processing message {i}: {sender} - {subject}")

                    for attachment in message.Attachments:
                        attachment_name = str(attachment.FileName).lower().strip()
                        matched = False

                        for match in possible_matches:
                            if fnmatch.fnmatch(attachment_name, match['AttachmentPattern']):
                                safe_name = re.sub(r'[<>:"/\\|?*]', '_', match['SaveName'])
                                save_path = os.path.join(self.archive_folder, safe_name)
                                attachment.SaveAsFile(save_path)

                                logger.info(f"Saved: {save_path}")

                                updates[match['RowIndex']] = {
                                    'Status': 'Saved'
                                }

                                matched = True

                        if not matched:
                            logger.info(f"No match for: {sender}, {subject}, {attachment_name}")

            except Exception as emsg_error:
                logger.error(f"Error processing message {i}: {emsg_error}")

        self.excel_doc_updates = updates

        logger.info(f"Emails Processed, Total attachments saved: {len(self.excel_doc_updates)}")

        return updates


    def update_excel_status(self, script_end, duration):
        if  self.excel_doc_updates is None:
            raise ValueError("Attachments haven't been saved, call match_and_save_attachments first")

        for idx, update in self.excel_doc_updates.items():
            excel_row = idx + 2  # +2 because Excel is 1-indexed and row 1 is header
            if 'Status' in update:
                status_col = self.mapping_df.columns.get_loc('Status') + 1
                self.auto_ws.cell(row=excel_row, column=status_col).value = update['Status']

        logger.info("Excel status and comments updated successfully.")

        named_cells = RestatementProcessor._get_named_cell_value(self.wb)

        named_cells["End_Time"].value = script_end.strftime("%Y-%m-%d %H:%M:%S")
        named_cells["Execution_Time"].value = str(duration).split('.')[0]

        try:
            self.wb.save(self.excel_file)
            logger.info(f"Workbook saved successfully to {self.excel_file}.")
        except Exception as msg_error:
            logger.error(f"Failed to save workbook: {msg_error}")


    def cleanup(self):
        try:
            # Release Outlook COM objects
            self.folder = None
            self.items = None

            try:
                pythoncom.CoUninitialize()
            except Exception as com_error:
                logger.error("Error during COM uninitialization: %s", com_error)

            # Release workbook and DataFrame
            self.wb = None
            self.auto_ws = None
            self.mapping_df = None
            self.mapping_dict = None

            gc.collect()
            logger.info("Processor cleanup completed.")
        except Exception as cleanup_error:
            logger.error("Error during processor cleanup: %s", cleanup_error)

# --------------------------------------------------- Main Function ----------------------------------------------------

def main():
    processor = None
    try:
        # Grabs dates and creates logs
        script_start = datetime.now()
        cbd = get_prior_business_days(0)
        pbd = get_prior_business_days(1)
        p2bd = get_prior_business_days(2)

        save_dir = make_import_archive_path(pbd)
        log_file_path = os.path.join(save_dir, "script.log")

        for handler in logging.root.handlers[:]:
            logging.root.removeHandler(handler)
        logging.basicConfig(filename=log_file_path, level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

        # Initialize COM for Outlook interaction
        pythoncom.CoInitialize()

        logger.info("Script execution started.")

        processor = RestatementProcessor(EXCEL_FILE, SHARED_MAILBOX, save_dir)
        processor.reset_excel_template(cbd, pbd, p2bd, script_start)
        processor.build_dictionary_from_excel()
        processor.connect_outlook(folder_name=FOLDER_NAME, folder_type=OutlookFolderName.CUSTOM)
        processor.get_items(get_filter_time(pbd, hour=16, minute=0, timezone="US/Eastern"))
        processor.match_and_save_attachments()

        script_end = datetime.now()
        duration = script_end - script_start
        logger.info(f"Script duration: {duration}")

        processor.update_excel_status(script_end, duration)

        # Summary output
        logger.info(f"Processed {len(processor.items)} messages.")
        logger.info(f"Saved {len(processor.excel_doc_updates)} attachments.")

        print(f"Processed {len(processor.items)} messages. Saved {len(processor.excel_doc_updates)} attachments.")

    except Exception as e:
        error_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        error_message = f"Unhandled exception occurred at {error_time}: {type(e).__name__} - {e}"
        logger.error(error_message)
        logger.exception("An unexpected error occurred.")
        sys.exit(1)

    finally:
        logger.info("Starting cleanup...")

        if 'processor' is not None:
            processor.cleanup()

# ------------------------------------------------- Main Function Call -------------------------------------------------

if __name__ == '__main__':
    main()